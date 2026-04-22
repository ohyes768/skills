#!/usr/bin/env python3
"""
从中国人民银行官网直接下载金融机构信贷收支表（xlsx），
解析中长期贷款数据。
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

# ── 本地 fetch_common（永远优先）─────────────────────────────────────────────
_SCRIPT_DIR = Path(__file__).resolve().parents[0]
_LOCAL_COMMON = _SCRIPT_DIR / "fetch_common.py"
_MONETARY_COMMON = (
    Path(__file__).resolve().parents[2]
    / "monetary-policy-skill"
    / "scripts"
    / "fetch_common.py"
)

# 确保本地路径在最前（先去重，再插到 0 位）
for _p in [str(_SCRIPT_DIR), str(_MONETARY_COMMON.parent)]:
    if _p in sys.path:
        sys.path.remove(_p)
sys.path.insert(0, str(_SCRIPT_DIR))  # 本地永远优先

from fetch_common import build_session, setup_logging, to_iso_now, write_cache, read_cache, is_data_published, LOGGER

PBOC_BASE = "https://www.pbc.gov.cn"


def _build_index_url(year: int) -> str:
    """动态生成 PBOC 索引页 URL（年份会自动更新）。"""
    return (
        f"{PBOC_BASE}/diaochatongjisi/116219/116319/{year}ntjsj/jrjgxdsztj/index.html"
    )


def get_current_index_url() -> str:
    """获取当前年份的 PBOC 索引页 URL，尝试最近3年以防当年页面未上线。"""
    current_year = datetime.now().year
    for year in range(current_year, current_year - 3, -1):
        url = _build_index_url(year)
        LOGGER.info("尝试 PBOC 索引页: %s", url)
        # 用 HEAD 请求验证页面是否存在
        try:
            session_test = build_session_with_retry()
            r = session_test.head(url, timeout=10)
            if r.status_code == 200:
                LOGGER.info("命中年份: %d", year)
                return url
        except Exception:
            pass
    # 兜底：返回今年 URL（页面可能未更新，但值得一试）
    return _build_index_url(current_year)


def build_session_with_retry() -> requests.Session:
    session = requests.Session()
    retries = Retry(total=3, connect=3, read=3, backoff_factor=0.6,
                    status_forcelist=(429, 500, 502, 503, 504))
    adapter = HTTPAdapter(max_retries=retries)
    session.mount("http://", adapter)
    session.mount("https://", adapter)
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/123.0.0.0",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "zh-CN,zh;q=0.9",
    })
    return session


def parse_pbc_page(session: requests.Session, index_url: str) -> list[dict[str, str]]:
    """解析 PBOC 索引页，提取所有 xlsx 下载链接"""
    r = session.get(index_url, timeout=20)
    # 页面编码混乱，用 GBK 解码（PBOC 官方用 GBK）
    r.encoding = "gbk"
    text = r.text

    # 找所有 .xlsx 链接
    links: list[dict[str, str]] = []
    pattern = re.compile(r'href="(/diaochatongjisi/attachDir/[^"]+\.xlsx)"')
    for m in pattern.finditer(text):
        url = PBOC_BASE + m.group(1)
        # 从 URL 提取标题（文件名中含日期）
        fname = m.group(1).split("/")[-1]
        # 找中文标题（在链接文本中）
        # 简单处理：直接用文件名的日期部分
        links.append({"url": url, "filename": fname})

    return links


def download_xlsx(session: requests.Session, url: str, dest: Path) -> bool:
    """下载 xlsx 文件到本地"""
    r = session.get(url, timeout=30)
    if r.status_code != 200:
        LOGGER.warning("下载失败 %s: %d", url, r.status_code)
        return False
    dest.write_bytes(r.content)
    LOGGER.info("已下载 %s -> %s", url, dest.name)
    return True


def parse_credit_balance_xlsx(xlsx_path: Path) -> dict[str, Any]:
    """解析金融机构本外币信贷收支 xlsx"""
    result: dict[str, Any] = {
        "parse_status": "failed",
        "data": {},
    }

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    # 找月份行（Row 7: 项目 Item | 2026.01 | 2026.02 | ...）
    month_cols: dict[str, int] = {}  # month_str -> col_index (0-based)

    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=7, column=col_idx).value
        if val is None:
            continue
        val_str = str(val).strip()
        # 匹配 "2026.01" / "2026.1" / "2025.01" / "2025.1" 格式（11月=11, 12月=12）
        m = re.match(r"(20\d{2})\.(\d+)", val_str)
        if m:
            year = int(m.group(1))
            month_num = int(m.group(2))
            if 1 <= month_num <= 12:
                month_cols[f"{year}-{month_num:02d}"] = col_idx - 1  # 0-based

    LOGGER.info("发现月份列: %s", list(month_cols.keys()))

    # 找关键行（Row 9 开始是数据）
    row_data: dict[str, dict[str, float | None]] = {}

    for row_idx in range(9, ws.max_row + 1):
        cell0 = ws.cell(row=row_idx, column=1).value
        label = str(cell0).strip() if cell0 else ""

        # 匹配中长期贷款行
        is_mlt_enterprise = "中长期" in label and "企业" in label
        is_mlt_household = "中长期" in label and "消费" in label
        is_mlt_total = "中长期" in label and "企业" not in label and "消费" not in label

        if is_mlt_enterprise or is_mlt_household or is_mlt_total:
            # 提取各月数据
            monthly: dict[str, float | None] = {}
            for month, col_0based in month_cols.items():
                val = ws.cell(row=row_idx, column=col_0based + 1).value  # openpyxl is 1-based
                if isinstance(val, (int, float)):
                    monthly[month] = float(val)
                else:
                    monthly[month] = None

            key = "中长期企业贷款" if is_mlt_enterprise else \
                  "中长期消费贷款" if is_mlt_household else "中长期贷款合计"
            row_data[key] = monthly
            LOGGER.info("找到行 [%s]: %s", key, {k: v for k, v in monthly.items() if v})

    if not row_data:
        result["error"] = "未找到中长期贷款行"
        return result

    result["data"] = row_data
    result["parse_status"] = "ok"
    return result


def fetch_pbc_credit_balance(target_month: str) -> dict[str, Any]:
    """下载并解析 PBOC 金融机构信贷收支表。

    PBOC 发布时间：每月15-20号左右发布上月数据
    """
    # 优先读缓存，避免重复下载 PBOC xlsx
    cached = read_cache("pbc_credit", target_month)
    if cached:
        LOGGER.info("缓存命中: pbc_credit/%s，直接返回", target_month)
        return cached

    # 检查数据是否可能已发布（PBOC 每月20号发布）
    published, hint = is_data_published(target_month, publish_day=20)
    if not published:
        LOGGER.info("PBOC 贷款数据 %s 尚未到发布时间（预计 %s）", target_month, hint)
        return {
            "target_month": target_month,
            "parse_status": "not_yet_published",
            "published_hint": hint,
            "fetched_at": to_iso_now(),
        }

    result: dict[str, Any] = {
        "target_month": target_month,
        "unit": "亿元",
        "source_url": "",
        "provider": "pbc.gov.cn",
        "parse_status": "failed",
    }

    session = build_session_with_retry()

    # 动态获取当前年份的 PBOC 索引页 URL
    index_url = get_current_index_url()
    result["source_url"] = index_url

    # 解析索引页找 xlsx 链接
    links = parse_pbc_page(session, index_url)
    if not links:
        result["error"] = "未找到 xlsx 下载链接"
        return result

    LOGGER.info("发现 %d 个 xlsx 文件", len(links))

    # 下载第一个（最完整）的 xlsx - 本外币信贷收支
    xlsx_info = next((l for l in links if "2026041416541518969" in l["url"]), links[0])
    xlsx_path = Path(tempfile.gettempdir()) / "pbc_credit_balance.xlsx"

    if not download_xlsx(session, xlsx_info["url"], xlsx_path):
        result["error"] = "xlsx 下载失败"
        return result

    # 解析
    parsed = parse_credit_balance_xlsx(xlsx_path)
    if parsed["parse_status"] != "ok":
        result["error"] = parsed.get("error", "解析失败")
        return result

    # 提取目标月份数据
    data = parsed["data"]
    row_keys = ["中长期企业贷款", "中长期消费贷款", "中长期贷款合计"]

    for key in row_keys:
        if key in data:
            val = data[key].get(target_month)
            result[key] = val

    result["parse_status"] = "ok"
    LOGGER.info("PBOC 数据提取成功: %s", {k: v for k, v in result.items() if k not in ["source_url", "provider"]})

    # 写入月度缓存
    write_cache("pbc_credit", target_month, result)

    return result


def main() -> None:
    parser = argparse.ArgumentParser(description="从 PBOC 官网下载中长期贷款数据")
    parser.add_argument("--month", required=True, help="目标月份（YYYY-MM）")
    parser.add_argument("--output", type=str, default="", help="输出 JSON 文件路径")
    args = parser.parse_args()

    setup_logging()
    data = fetch_pbc_credit_balance(args.month)
    rendered = json.dumps(data, ensure_ascii=False, indent=2)
    print(rendered)

    if args.output:
        Path(args.output).write_text(rendered + "\n", encoding="utf-8")
        LOGGER.info("已写入 %s", args.output)


if __name__ == "__main__":
    main()